"""
Excel Trend Analyzer — Temperature & Humidity
- Win32를 이용해 Excel 파일 오픈
- 전체 시트 각각에 독립적으로 추세 분석 적용
- 온도 / 습도 각각 독립된 min_rows, min_val 설정 (config.json)
- 시트별 시간 / 온도 / 습도 컬럼 개별 지정
- 온도 → Temp_Trend, 습도 → Humid_Trend 컬럼 추가
- 노이즈 구간은 앞뒤 추세를 반영하여 채움
"""

import json
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import win32com.client
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment


# ──────────────────────────────────────────────
# 설정 파일
# ──────────────────────────────────────────────
CONFIG_FILE = "config.json"

DEFAULT_CONFIG = {
    "temp":  {"min_rows": 3, "min_val": 0.0},
    "humid": {"min_rows": 3, "min_val": 0.0},
}


def load_config() -> dict:
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            raw = json.load(f)
        # 누락 키 보완
        cfg = {}
        for key in ("temp", "humid"):
            base = DEFAULT_CONFIG[key].copy()
            base.update(raw.get(key, {}))
            cfg[key] = base
        return cfg
    return {k: v.copy() for k, v in DEFAULT_CONFIG.items()}


def save_config(cfg: dict):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)


# ──────────────────────────────────────────────
# 추세 분석 로직 (공통)
# ──────────────────────────────────────────────
def analyze_trends(values: list, min_rows: int, min_val: float) -> list:
    n = len(values)
    if n == 0:
        return []

    raw = [""] * n
    for i in range(1, n):
        if values[i] > values[i - 1]:
            raw[i] = "UP"
        elif values[i] < values[i - 1]:
            raw[i] = "DOWN"

    for i in range(1, n):
        if raw[i] == "":
            raw[i] = raw[i - 1]

    segments = []
    i = 0
    while i < n:
        d = raw[i]; j = i
        while j < n and raw[j] == d:
            j += 1
        segments.append((i, j - 1, d))
        i = j

    low_set = {k for k, v in enumerate(values) if v <= min_val}

    def is_noise(si):
        s, e, _ = segments[si]
        return all(k in low_set for k in range(s, e + 1)) or (e - s + 1) < min_rows

    resolved = [None] * len(segments)
    for si in range(len(segments)):
        if not is_noise(si):
            resolved[si] = segments[si][2]

    last = ""
    for si in range(len(segments)):
        if resolved[si] is not None:
            last = resolved[si]
        else:
            resolved[si] = last

    nxt = ""
    for si in range(len(segments) - 1, -1, -1):
        if not is_noise(si):
            nxt = segments[si][2]
        elif resolved[si] == "":
            resolved[si] = nxt

    result = [""] * n
    for si, (s, e, _) in enumerate(segments):
        for k in range(s, e + 1):
            result[k] = resolved[si] or ""
    return result


# ──────────────────────────────────────────────
# 파일 / 시트 유틸
# ──────────────────────────────────────────────
def get_all_sheet_names(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    names = wb.sheetnames; wb.close()
    return names


def get_sheet_headers(filepath, sheet_name):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    wb.close()
    return [str(h) for h in headers if h is not None]


# ──────────────────────────────────────────────
# 스타일 상수
# ──────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", start_color="2E4057")
HEADER_FONT = Font(bold=True, color="FFFFFF")

T_UP_FILL   = PatternFill("solid", start_color="C8E6C9")
T_DOWN_FILL = PatternFill("solid", start_color="FFCDD2")
T_UP_FONT   = Font(color="1B5E20", bold=True)
T_DOWN_FONT = Font(color="B71C1C", bold=True)

H_UP_FILL   = PatternFill("solid", start_color="B3E5FC")
H_DOWN_FILL = PatternFill("solid", start_color="FFE0B2")
H_UP_FONT   = Font(color="01579B", bold=True)
H_DOWN_FONT = Font(color="E65100", bold=True)

CENTER = Alignment(horizontal="center")


# ──────────────────────────────────────────────
# 단일 컬럼 추세 기록
# ──────────────────────────────────────────────
def write_trend_column(ws_read, ws_write, src_col, result_col_name,
                       min_rows, min_val, up_fill, down_fill, up_font, down_font):
    headers = [str(c.value) if c.value is not None else "" for c in ws_read[1]]
    if src_col not in headers:
        return None

    src_idx = headers.index(src_col)
    values = []
    for row in ws_read.iter_rows(min_row=2, values_only=True):
        val = row[src_idx] if src_idx < len(row) else None
        try:
            values.append(float(val) if val is not None else 0.0)
        except (TypeError, ValueError):
            values.append(0.0)

    trends = analyze_trends(values, min_rows, min_val)

    existing = [ws_write.cell(row=1, column=c).value
                for c in range(1, ws_write.max_column + 1)]
    col_idx = (existing.index(result_col_name) + 1
               if result_col_name in existing
               else ws_write.max_column + 1)

    hc = ws_write.cell(row=1, column=col_idx, value=result_col_name)
    hc.font = HEADER_FONT; hc.fill = HEADER_FILL; hc.alignment = CENTER

    for i, trend in enumerate(trends):
        cell = ws_write.cell(row=i + 2, column=col_idx, value=trend)
        cell.alignment = CENTER
        if trend == "UP":
            cell.fill = up_fill; cell.font = up_font
        elif trend == "DOWN":
            cell.fill = down_fill; cell.font = down_font

    ws_write.column_dimensions[
        ws_write.cell(row=1, column=col_idx).column_letter].width = 14
    return trends


# ──────────────────────────────────────────────
# 전체 파일 처리
# ──────────────────────────────────────────────
def process_all_sheets(filepath, sheet_cfg_map, config, status_cb=None):
    def log(msg):
        if status_cb: status_cb(msg)

    t_cfg = config["temp"]
    h_cfg = config["humid"]

    log("Excel 파일 읽는 중...")
    wb_read  = openpyxl.load_workbook(filepath, data_only=True)
    wb_write = openpyxl.load_workbook(filepath)
    processed = []

    for sheet_name, scfg in sheet_cfg_map.items():
        if not scfg.get("enabled", True):
            log(f"  [{sheet_name}] 스킵 (비활성)")
            continue

        ws_r = wb_read[sheet_name]
        ws_w = wb_write[sheet_name]
        t_col = scfg.get("temp_col",  "")
        h_col = scfg.get("humid_col", "")
        sheet_done = False

        if t_col:
            trends = write_trend_column(
                ws_r, ws_w, t_col, "Temp_Trend",
                int(t_cfg["min_rows"]), float(t_cfg["min_val"]),
                T_UP_FILL, T_DOWN_FILL, T_UP_FONT, T_DOWN_FONT)
            if trends is not None:
                log(f"  [{sheet_name}] 🌡 온도({t_col})  "
                    f"min_rows={t_cfg['min_rows']}, min_val={t_cfg['min_val']}  "
                    f"→ UP={trends.count('UP')}, DOWN={trends.count('DOWN')}")
                sheet_done = True
            else:
                log(f"  [{sheet_name}] 🌡 온도 컬럼 '{t_col}' 없음 → 스킵")

        if h_col:
            trends = write_trend_column(
                ws_r, ws_w, h_col, "Humid_Trend",
                int(h_cfg["min_rows"]), float(h_cfg["min_val"]),
                H_UP_FILL, H_DOWN_FILL, H_UP_FONT, H_DOWN_FONT)
            if trends is not None:
                log(f"  [{sheet_name}] 💧 습도({h_col})  "
                    f"min_rows={h_cfg['min_rows']}, min_val={h_cfg['min_val']}  "
                    f"→ UP={trends.count('UP')}, DOWN={trends.count('DOWN')}")
                sheet_done = True
            else:
                log(f"  [{sheet_name}] 💧 습도 컬럼 '{h_col}' 없음 → 스킵")

        if not t_col and not h_col:
            log(f"  [{sheet_name}] 컬럼 미설정 → 스킵")
        elif sheet_done:
            processed.append(sheet_name)

    base, ext = os.path.splitext(filepath)
    out_path = base + "_trend" + ext
    wb_write.save(out_path)
    log(f"\n저장: {out_path}")
    log(f"처리 완료 시트: {len(processed)}개")

    log("Excel에서 파일 여는 중...")
    try:
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Visible = True
        xl.Workbooks.Open(os.path.abspath(out_path))
        log("Excel 열기 완료!")
    except Exception as e:
        log(f"[경고] Win32 Excel 열기 실패: {e}")

    return out_path, processed


# ──────────────────────────────────────────────
# GUI: 시트 1행 위젯
# ──────────────────────────────────────────────
class SheetRow:
    def __init__(self, parent, sheet_name, headers, bg, fg, entry_bg, lf, row_idx):
        self.sheet_name = sheet_name
        self.enabled   = tk.BooleanVar(value=True)
        self.time_var  = tk.StringVar()
        self.temp_var  = tk.StringVar()
        self.humid_var = tk.StringVar()
        NONE_OPT = ["(없음)"] + headers

        tk.Checkbutton(parent, variable=self.enabled, bg=bg, fg=fg,
                       selectcolor=entry_bg, activebackground=bg,
                       command=self._toggle).grid(row=row_idx, column=0, padx=(6, 2))

        tk.Label(parent, text=sheet_name, font=lf, bg=bg, fg=fg,
                 width=12, anchor="w").grid(row=row_idx, column=1, padx=4)

        self.time_cb = ttk.Combobox(parent, textvariable=self.time_var,
                                    values=headers, width=14, state="readonly")
        self.time_cb.grid(row=row_idx, column=2, padx=3, pady=2)
        if headers: self.time_cb.current(0)

        self.temp_cb = ttk.Combobox(parent, textvariable=self.temp_var,
                                    values=NONE_OPT, width=14, state="readonly")
        self.temp_cb.grid(row=row_idx, column=3, padx=3, pady=2)
        self.temp_cb.current(2 if len(headers) >= 2 else 0)

        self.humid_cb = ttk.Combobox(parent, textvariable=self.humid_var,
                                     values=NONE_OPT, width=14, state="readonly")
        self.humid_cb.grid(row=row_idx, column=4, padx=3, pady=2)
        self.humid_cb.current(3 if len(headers) >= 3 else 0)

        self._combos = [self.time_cb, self.temp_cb, self.humid_cb]

    def _toggle(self):
        state = "readonly" if self.enabled.get() else "disabled"
        for cb in self._combos: cb.configure(state=state)

    def get(self):
        tc = self.temp_var.get(); hc = self.humid_var.get()
        return {
            "enabled":   self.enabled.get(),
            "time_col":  self.time_var.get(),
            "temp_col":  "" if tc in ("", "(없음)") else tc,
            "humid_col": "" if hc in ("", "(없음)") else hc,
        }


# ──────────────────────────────────────────────
# GUI: 설정 그룹 (min_rows + min_val 한 세트)
# ──────────────────────────────────────────────
class SensorConfigFrame(tk.LabelFrame):
    """온도 또는 습도의 min_rows / min_val 입력 그룹"""

    def __init__(self, parent, label, init: dict, bg, fg, entry_bg, lf):
        super().__init__(parent, text=label, font=("Segoe UI", 10, "bold"),
                         bg=bg, fg=fg, bd=1, relief="groove",
                         labelanchor="nw", padx=8, pady=6)
        self.configure(background=bg)

        self.min_rows_var = tk.StringVar(value=str(init.get("min_rows", 3)))
        self.min_val_var  = tk.StringVar(value=str(init.get("min_val",  0.0)))

        for i, (lbl, var) in enumerate([
            ("min_rows  (노이즈 최소 행)", self.min_rows_var),
            ("min_val   (노이즈 임계값)",  self.min_val_var),
        ]):
            tk.Label(self, text=lbl, font=lf, bg=bg, fg=fg).grid(
                row=i, column=0, sticky="e", padx=(0, 6), pady=3)
            tk.Entry(self, textvariable=var, width=10,
                     bg=entry_bg, fg=fg, insertbackground=fg,
                     relief="flat").grid(row=i, column=1, sticky="w", pady=3)

    def get(self) -> dict:
        return {
            "min_rows": int(self.min_rows_var.get()),
            "min_val":  float(self.min_val_var.get()),
        }

    def set(self, d: dict):
        self.min_rows_var.set(str(d.get("min_rows", 3)))
        self.min_val_var.set(str(d.get("min_val",  0.0)))


# ──────────────────────────────────────────────
# GUI: 메인 앱
# ──────────────────────────────────────────────
class App(tk.Tk):
    BG  = "#1E1E2E";  FG  = "#CDD6F4"
    EBG = "#313244";  BBG = "#89B4FA";  BFG = "#1E1E2E"
    SEP = "#45475A"
    LF  = ("Segoe UI", 10)
    TF  = ("Segoe UI", 13, "bold")
    PAD = 12

    def __init__(self):
        super().__init__()
        self.title("Trend Analyzer — Temperature & Humidity")
        self.configure(bg=self.BG)
        self.resizable(True, False)
        self.config_data = load_config()
        self._sheet_rows: list[SheetRow] = []
        self._build_ui()

    def _build_ui(self):
        P = self.PAD

        tk.Label(self, text="🌡💧 Temperature & Humidity Trend Analyzer",
                 font=self.TF, bg=self.BG, fg=self.BBG).pack(pady=(P, 4))

        # 파일 선택
        ff = tk.Frame(self, bg=self.BG)
        ff.pack(fill="x", padx=P, pady=4)
        tk.Label(ff, text="Excel 파일", font=self.LF,
                 bg=self.BG, fg=self.FG, width=10, anchor="e").pack(side="left")
        self.file_var = tk.StringVar()
        tk.Entry(ff, textvariable=self.file_var, bg=self.EBG, fg=self.FG,
                 insertbackground=self.FG, relief="flat").pack(
            side="left", fill="x", expand=True, padx=6)
        tk.Button(ff, text="찾아보기", bg=self.BBG, fg=self.BFG,
                  font=self.LF, relief="flat", cursor="hand2",
                  command=self._browse).pack(side="left")

        self._sep()

        # 시트 테이블 헤더
        tk.Label(self, text="📋 시트별 컬럼 설정",
                 font=("Segoe UI", 10, "bold"),
                 bg=self.BG, fg="#A6E3A1").pack(pady=(0, 2))

        hf = tk.Frame(self, bg=self.BG)
        hf.pack(fill="x", padx=P)
        for col, txt, w in [
            (0,"적용",4),(1,"시트명",12),(2,"시간 컬럼",15),
            (3,"🌡 온도 컬럼",15),(4,"💧 습도 컬럼",15)
        ]:
            tk.Label(hf, text=txt, font=("Segoe UI", 9, "bold"),
                     bg="#313244", fg="#A6ADC8", width=w,
                     padx=4, pady=2).grid(row=0, column=col, padx=2, pady=(0, 2))

        # 스크롤 캔버스
        self.canvas = tk.Canvas(self, bg=self.BG, highlightthickness=0, height=200)
        vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=vsb.set)
        self.canvas.pack(side="left", fill="both", expand=True, padx=(P, 0))
        vsb.pack(side="left", fill="y", padx=(0, P))

        self.sf = tk.Frame(self.canvas, bg=self.BG)
        self._cw = self.canvas.create_window((0, 0), window=self.sf, anchor="nw")
        self.sf.bind("<Configure>",
                     lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.bind("<Configure>",
                         lambda e: self.canvas.itemconfig(self._cw, width=e.width))

        tk.Label(self.sf, text="← Excel 파일을 선택하면 시트 목록이 표시됩니다",
                 font=self.LF, bg=self.BG, fg="#585B70").grid(
            row=0, column=0, columnspan=5, pady=20)

        self._sep()

        # ── 설정 (온도 / 습도 나란히) ──────────────
        tk.Label(self, text="⚙ 설정 (config.json)",
                 font=("Segoe UI", 10, "bold"),
                 bg=self.BG, fg="#A6E3A1").pack(pady=(0, 4))

        cfg_row = tk.Frame(self, bg=self.BG)
        cfg_row.pack(padx=P, pady=4)

        self.temp_cfg_frame = SensorConfigFrame(
            cfg_row, "🌡 온도 (temp)",
            self.config_data["temp"],
            self.BG, self.FG, self.EBG, self.LF)
        self.temp_cfg_frame.grid(row=0, column=0, padx=(0, 12), sticky="n")

        self.humid_cfg_frame = SensorConfigFrame(
            cfg_row, "💧 습도 (humid)",
            self.config_data["humid"],
            self.BG, self.FG, self.EBG, self.LF)
        self.humid_cfg_frame.grid(row=0, column=1, sticky="n")

        tk.Button(cfg_row, text="설정\n저장", bg="#A6E3A1", fg=self.BFG,
                  font=self.LF, relief="flat", cursor="hand2",
                  command=self._save_cfg).grid(row=0, column=2, padx=(12, 0),
                                               ipadx=6, ipady=4, sticky="ns")

        self._sep()

        tk.Button(self, text="▶  전체 시트 분석 실행",
                  bg=self.BBG, fg=self.BFG,
                  font=("Segoe UI", 11, "bold"), relief="flat", cursor="hand2",
                  command=self._run).pack(pady=(0, 6), ipadx=20, ipady=4)

        self.log_box = tk.Text(self, height=8, bg="#181825", fg="#BAC2DE",
                               font=("Consolas", 9), relief="flat", state="disabled")
        self.log_box.pack(fill="x", padx=P, pady=(0, P))

    def _sep(self):
        tk.Frame(self, bg=self.SEP, height=1).pack(fill="x", padx=self.PAD, pady=6)

    # ── 이벤트 ────────────────────────────────────
    def _browse(self):
        path = filedialog.askopenfilename(
            title="Excel 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")])
        if not path: return
        self.file_var.set(path)
        self._load_sheets(path)

    def _load_sheets(self, path):
        for w in self.sf.winfo_children(): w.destroy()
        self._sheet_rows.clear()
        try:
            names = get_all_sheet_names(path)
        except Exception as e:
            messagebox.showerror("오류", f"파일 읽기 실패:\n{e}"); return

        for i, sname in enumerate(names):
            try:
                hdrs = get_sheet_headers(path, sname)
            except Exception:
                hdrs = []
            self._sheet_rows.append(
                SheetRow(self.sf, sname, hdrs,
                         self.BG, self.FG, self.EBG, self.LF, i))

        self._log(f"파일 로드 완료: {len(names)}개 시트 — " + ", ".join(names))

    def _save_cfg(self):
        try:
            cfg = {
                "temp":  self.temp_cfg_frame.get(),
                "humid": self.humid_cfg_frame.get(),
            }
            save_config(cfg)
            self.config_data = cfg
            self._log(
                f"설정 저장 완료\n"
                f"  🌡 temp : min_rows={cfg['temp']['min_rows']}, "
                f"min_val={cfg['temp']['min_val']}\n"
                f"  💧 humid: min_rows={cfg['humid']['min_rows']}, "
                f"min_val={cfg['humid']['min_val']}")
        except ValueError:
            messagebox.showerror("오류", "min_rows는 정수, min_val은 실수로 입력하세요.")

    def _run(self):
        filepath = self.file_var.get().strip()
        if not filepath or not os.path.exists(filepath):
            messagebox.showerror("오류", "유효한 Excel 파일을 선택하세요."); return
        if not self._sheet_rows:
            messagebox.showerror("오류", "시트 정보가 없습니다. 파일을 다시 선택하세요."); return

        try:
            cfg = {
                "temp":  self.temp_cfg_frame.get(),
                "humid": self.humid_cfg_frame.get(),
            }
        except ValueError:
            messagebox.showerror("오류", "min_rows / min_val 값을 확인하세요."); return

        sheet_cfg_map = {sr.sheet_name: sr.get() for sr in self._sheet_rows}

        self._log("═" * 56)
        self._log(f"파일: {os.path.basename(filepath)}")
        self._log(f"🌡 temp  min_rows={cfg['temp']['min_rows']}, "
                  f"min_val={cfg['temp']['min_val']}")
        self._log(f"💧 humid min_rows={cfg['humid']['min_rows']}, "
                  f"min_val={cfg['humid']['min_val']}")

        def status(msg):
            self._log(msg); self.update_idletasks()

        try:
            out_path, done = process_all_sheets(
                filepath, sheet_cfg_map, cfg, status_cb=status)
            messagebox.showinfo(
                "완료",
                f"분석 완료!\n처리 시트: {len(done)}개\n저장: {os.path.basename(out_path)}")
        except Exception as e:
            self._log(f"[오류] {e}"); messagebox.showerror("오류", str(e))

    def _log(self, msg):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg + "\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")


# ──────────────────────────────────────────────
if __name__ == "__main__":
    app = App()
    app.mainloop()
